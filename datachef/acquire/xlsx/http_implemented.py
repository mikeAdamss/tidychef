"""
Holds the code that defines the local xlsx reader.
"""

import copy
import io
from pathlib import Path
from typing import Any, Callable, List, Optional, Union

import openpyxl
import requests

from datachef.acquire.base import BaseReader
from datachef.models.source.cell import Cell
from datachef.models.source.table import Table
from datachef.selection.selectable import Selectable
from datachef.selection.xlsx.xlsx import XlsxSelectable
from datachef.utils.http.caching import get_cached_session

from ..base import BaseReader
from ..main import acquirer


def http(
    source: Union[str, Path],
    selectable: Selectable = XlsxSelectable,
    pre_hook: Optional[Callable] = None,
    post_hook: Optional[Callable] = None,
    session: requests.Session = None,
    cache: bool = True,
    **kwargs,
) -> List[XlsxSelectable]:
    """
    Read data from a Path (or string representing a path)
    present on the same machine where datachef is running.

    This local xlsx reader uses openpyxl:
    https://openpyxl.readthedocs.io/en/stable/index.html

    Any kwargs passed to this function are propogated to
    the openpyxl.load_workbook() method
    """

    assert isinstance(
        source, (str, Path)
    ), """
        The source you're passing to acquire.csv.local() needs to
        be either a Path object or a string representing such.
        """

    return acquirer(
        source,
        HttpXlsxReader,
        selectable,
        pre_hook=pre_hook,
        post_hook=post_hook,
        session=session,
        cache=cache,
        **kwargs,
    )


class HttpXlsxReader(BaseReader):
    """
    A reader to lead in a source where that source is a locally
    held xlsx file.
    """

    def parse(
        source: Any,
        selectable: Selectable = XlsxSelectable,
        data_only=True,
        session: requests.Session = None,
        cache: bool = True,
        **kwargs,
    ) -> List[XlsxSelectable]:

        if cache:
            session = get_cached_session()
        else:
            session = requests.session()

        response: requests.Response = session.get(source, **kwargs)
        if not response.ok:
            raise requests.exceptions.HTTPError(
                f"""
                Unable to get url: {source}
                {response}
                """
            )

        bio = io.BytesIO()
        bio.write(response.content)
        bio.seek(0)

        workbook: openpyxl.Workbook = openpyxl.load_workbook(bio, data_only=data_only)

        datachef_selectables = []
        worksheet_names = workbook.get_sheet_names()
        for worksheet_name in worksheet_names:

            worksheet = workbook.get_sheet_by_name(worksheet_name)

            table = Table()
            for y, row in enumerate(worksheet.iter_rows()):
                for x, cell in enumerate(row):
                    table.add_cell(Cell(x=x, y=y, value=cell.value if cell.value else ""))

            datachef_selectables.append(
                selectable(
                    table, copy.deepcopy(table), source=source, _name=worksheet_name
                )
            )
        return datachef_selectables
