from typing import Any, Dict, List

from openpyxl.cell.cell import Cell as OpenPyxlCell
from openpyxl.workbook import Workbook

from tidychef.exceptions import UnknownExcelTimeError
from tidychef.models.source.cell import Cell
from tidychef.models.source.table import Table
from tidychef.selection import Selectable, XlsxSelectable

from ..excel_time import EXCEL_TIME_FORMATS, missing_time_format_message


def xlsx_time_formats():
    """
    Wrapped for test purposes
    """
    return EXCEL_TIME_FORMATS


def sheets_from_workbook(
    source: Any,
    selectable: Selectable,
    workbook: Workbook,
    custom_time_formats: Dict[str, str],
) -> List[XlsxSelectable]:
    tidychef_selectables = []
    for worksheet_name in workbook.sheetnames:

        worksheet = workbook[worksheet_name]

        # Unmerge all cells as merged cells do not have the
        # properties we need, i.e .is_date
        for items in sorted(worksheet.merged_cells.ranges): # pragma: no cover
            worksheet.unmerge_cells(str(items))

        table = Table()
        for y, row in enumerate(worksheet.iter_rows()):

            opycell: OpenPyxlCell
            for x, opycell in enumerate(row):
                if opycell.is_date and opycell.internal_value is not None:
                    strformat_pattern = xlsx_time_formats().get(
                        opycell.number_format, None
                    )
                    if strformat_pattern is None:
                        strformat_pattern = custom_time_formats.get(
                            opycell.number_format, None
                        )
                        if strformat_pattern is None:
                            raise UnknownExcelTimeError(
                                missing_time_format_message(opycell.number_format)
                            )
                    cell_value = opycell.internal_value.strftime(strformat_pattern)
                elif opycell.value is not None:
                    cell_value = opycell.value
                else:
                    cell_value = ""
                table.add_cell(Cell(x=x, y=y, value=str(cell_value)))

        tidychef_selectables.append(
            selectable(table, source=source, name=worksheet_name)
        )
    return tidychef_selectables
